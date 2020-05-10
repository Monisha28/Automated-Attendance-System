import face_recognition
import cv2
import numpy as np
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Color, Fill
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell


video_capture = cv2.VideoCapture(0)

moni = face_recognition.load_image_file("photos/20191110_170351.jpg")
moni_face_encoding = face_recognition.face_encodings(moni)[0]

# keerthana = face_recognition.load_image_file("photos/IMG-20191211-WA0010.jpg")
# keerthana_face_encoding = face_recognition.face_encodings(keerthana)[0]

# viji = face_recognition.load_image_file("/Users/moni/Downloads/20191214_123517.jpg")
# viji_face_encoding = face_recognition.face_encodings(viji)[0]

divya = face_recognition.load_image_file("photos/Divya Praba.jpg")
divya_face_encoding = face_recognition.face_encodings(divya)[0]

geetha = face_recognition.load_image_file("photos/Geethanjali.jpg")
geetha_face_encoding = face_recognition.face_encodings(geetha)[0]

goutham = face_recognition.load_image_file("photos/Goutham Raj R.jpg")
goutham_face_encoding = face_recognition.face_encodings(goutham)[0]

jayanth = face_recognition.load_image_file("photos/IMG-20191102-WA0039.jpg")
jayanth_face_encoding = face_recognition.face_encodings(jayanth)[0]

madhu = face_recognition.load_image_file("photos/madhu.jpg")
madhu_face_encoding = face_recognition.face_encodings(madhu)[0]

malathy = face_recognition.load_image_file("photos/Malathy.jpg")
malathy_face_encoding = face_recognition.face_encodings(malathy)[0]

meera = face_recognition.load_image_file("photos/Meera.png")
meera_face_encoding = face_recognition.face_encodings(meera)[0]

monica = face_recognition.load_image_file("photos/Monica Rachel .jpg")
monica_face_encoding = face_recognition.face_encodings(monica)[0]

nalini = face_recognition.load_image_file("photos/Nalini.jpg")
nalini_face_encoding = face_recognition.face_encodings(nalini)[0]

ruva = face_recognition.load_image_file("photos/Ruvanthika P A.jpg")
ruva_face_encoding = face_recognition.face_encodings(ruva)[0]

mithra = face_recognition.load_image_file("photos/Sangamithra Goutham.JPG")
mithra_face_encoding = face_recognition.face_encodings(mithra)[0]

sandy = face_recognition.load_image_file("photos/Santhiya C.jpeg")
sandy_face_encoding = face_recognition.face_encodings(sandy)[0]

vignesh = face_recognition.load_image_file("photos/Vignesh_Kumar.jpg")
vignesh_face_encoding = face_recognition.face_encodings(vignesh)[0]

known_face_encodings = [
    moni_face_encoding,
    divya_face_encoding,
    geetha_face_encoding,
    goutham_face_encoding,
    jayanth_face_encoding,
    madhu_face_encoding,
    malathy_face_encoding,
    monica_face_encoding,
    nalini_face_encoding,
    ruva_face_encoding,
    mithra_face_encoding,
    sandy_face_encoding,
    vignesh_face_encoding   
]
known_face_names = [
    "Monisha",
    "Divya",
    "Geethanjali",
    "Goutham",
    "Jayanth",
    "Madhu",
    "Malathy",
    "Monica",
    "Nalini",
    "Ruvanthika",
    "Sangamithra",
    "Santhiya",
    "Vignesh"
    ]
known_face_num = [
        
        "1",
        "2",
        "3",
        "4",
        "5",
        "6",
        "7",
        "8",
        "9",
        "10",
        "11",
        "12",
        "13"

]
# load xlsx
book=load_workbook('attendance.xlsx')

# load present date and time
now= datetime.datetime.now()
today=now.day
month=now.month

#create sheets
# sheets = book.sheetnames
# ws = book[sheets[1]]
# for row in ws['I1:I2']:
#   for cell in row:
#     cell.value = None
sheet=book.active
if sheet.title != now.strftime("%b(%m)"):
    sheet_name = now.strftime("%b(%m)")
    sheet = book.create_sheet(sheet_name, 0)
else:
    sheet=book.active
    
sheet.cell(row = 1, column = 1).value = 'Roll Number'
sheet.cell(row = 1, column = 2).value = 'Name'
i=2
j=2
for rollno in known_face_num:
    sheet.cell(row = i, column = 1).value = rollno
    i += 1  
for allname in known_face_names:
    sheet.cell(row = j, column = 2).value = allname
    j += 1  
for ws_column in range(1,34):
    col_letter = get_column_letter(ws_column)
    sheet.column_dimensions[col_letter].width = 16

face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    # caputure video from webcam and resize
    ret, frame = video_capture.read()

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    rgb_small_frame = small_frame[:, :, ::-1]

    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        
        face_names = []
        for face_encoding in face_encodings:
            matches = face_recognition.api.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            name = "Unknown"
            num = 0

            # If a match was found in known_face_encodings.
            if True in matches:
                # first_match_index = matches.index(True)
                # name = known_face_names[first_match_index]
                # num = known_face_num[first_match_index]

                # # Or instead, use the known face with the smallest distance to the new face
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    num = known_face_num[best_match_index]

            td = now.strftime("%d-%m-%Y (%a)")
            sheet.cell(row= 1, column=int(today)+2).value = td
            ab=1
            for ab in range(2,15):
                sheet.cell(row=ab, column=int(today)+2).value = "absent" 
                ab += 1
            # Assign attendance
            if int(num) in range(1,20):
                time = now.strftime("%H:%M:%S")
                sheet.cell(row=int(num)+1, column=int(today)+2).value = time  
            else:
                pass
                 
            face_names.append(name)

    process_this_frame = not process_this_frame

#style
    color_font = Font(color='0000FF00', italic=True, bold=True, name="Helvetica")
    style1 = Alignment(wrap_text=True)
    style2 = Font(name="Helvetica")

    for cell in sheet["1:1"]:
        cell.font = color_font 
       
    for row in sheet.iter_rows(min_row=1, max_col=34):
        for cell in row:
            cell.alignment = style1
            cell.font = style2
    
    
    process_this_frame = not process_this_frame

#display box with name
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    cv2.imshow('Video', frame)

    book.save('attendance.xlsx')
  
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()